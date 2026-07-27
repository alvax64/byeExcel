"""Builds the demo workbook: a Peruvian hardware store's real-life Excel.
Spanish headers with accents (exercises accent-insensitive relation matching),
three related sheets, formulas, blanks, and mixed types."""
from datetime import datetime
from openpyxl import Workbook

wb = Workbook()

ventas = wb.active
ventas.title = "Ventas"
ventas.append(["ID", "Fecha", "Cliente", "Producto", "Cantidad", "Precio Unitario", "Total", "Pagado"])
rows = [
    (1, datetime(2026, 7, 1), "Ana Quispe", "Tornillo 3in", 100, 0.15, "=E2*F2", True),
    (2, datetime(2026, 7, 3), "Luis Mamani", "Martillo", 2, 35.0, "=E3*F3", True),
    (3, datetime(2026, 7, 8), "Ana Quispe", "Clavo 2in", 500, 0.08, "=E4*F4", False),
    (4, datetime(2026, 7, 12), "Rosa Huamán", "Taladro Bosch", 1, 289.9, "=E5*F5", True),
    (5, datetime(2026, 7, 15), "Luis Mamani", "Tornillo 3in", 250, 0.15, "=E6*F6", False),
    (6, datetime(2026, 7, 20), "Carlos Ríos", "Sierra manual", 1, 45.5, "=E7*F7", True),
    (7, datetime(2026, 7, 22), "Rosa Huamán", "Clavo 2in", 200, 0.08, "=E8*F8", True),
    (8, datetime(2026, 7, 25), "Ana Quispe", "Martillo", 1, 35.0, "=E9*F9", None),
]
for r in rows:
    ventas.append(list(r))

clientes = wb.create_sheet("Clientes")
clientes.append(["DNI", "Nombre", "Teléfono", "Distrito"])
for r in [
    ("42715983", "Ana Quispe", "987654321", "San Juan de Lurigancho"),
    ("40118276", "Luis Mamani", "912345678", "Villa El Salvador"),
    ("45620134", "Rosa Huamán", "955443322", "Comas"),
    ("41893755", "Carlos Ríos", None, "Ate"),
]:
    clientes.append(list(r))

productos = wb.create_sheet("Productos")
productos.append(["SKU", "Producto", "Categoría", "Stock", "Precio"])
for r in [
    ("TOR-3", "Tornillo 3in", "Ferretería", 4500, 0.15),
    ("MAR-1", "Martillo", "Herramientas", 12, 35.0),
    ("CLA-2", "Clavo 2in", "Ferretería", 8000, 0.08),
    ("TAL-B", "Taladro Bosch", "Eléctricos", 3, 289.9),
    ("SIE-M", "Sierra manual", "Herramientas", 7, 45.5),
]:
    productos.append(list(r))

wb.save("demo/ferreteria_dona_rosa.xlsx")
print("wrote demo/ferreteria_dona_rosa.xlsx")
